"""Load Table"""
import os
import openpyxl
from selectFile import SF_var


def loading(line: int) -> list[list]:
    information_list_top: list[list] = []
    information_list_secondary: list = []
    for i in sheet.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=column):
        for j in i:
            information_list_secondary.append(j.value)
    for s in range(1, len(information_list_secondary) + 1):
        if s % line == 0:
            information_list_top.append(information_list_secondary[s - line:s])
    return information_list_top


if len(SF_var.get) != 0:
    path: str = SF_var.path_data
    os.path.abspath('.')
    path_s = os.path.split(path)
    workbook = openpyxl.load_workbook(path_s[1])
    sheet = workbook.active
    rows = sheet.max_row
    column = sheet.max_column
else:
    quit()
